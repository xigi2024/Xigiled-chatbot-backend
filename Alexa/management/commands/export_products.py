from django.core.management.base import BaseCommand
from Alexa.models import Product
import openpyxl
from openpyxl.styles import Font, Alignment
import os

class Command(BaseCommand):
    help = 'Export all products to an Excel file'

    def handle(self, *args, **options):
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Define headers
        headers = ['Name', 'Description', 'Price', 'Category', 'Guide Steps', 'Created At']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Get all products
        products = Product.objects.all()

        # Write product data
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1).value = product.name
            ws.cell(row=row_num, column=2).value = product.description
            ws.cell(row=row_num, column=3).value = product.price
            ws.cell(row=row_num, column=4).value = product.category or ''
            # Combine guide_steps into a single cell with line breaks
            guide_steps_text = '\n'.join(product.guide_steps) if product.guide_steps else ''
            ws.cell(row=row_num, column=5).value = guide_steps_text
            ws.cell(row=row_num, column=6).value = product.created_at.strftime('%Y-%m-%d %H:%M:%S')

            # Set alignment for guide_steps cell to wrap text
            ws.cell(row=row_num, column=5).alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # Cap at 50 for readability

        # Save the file
        file_path = os.path.join(os.getcwd(), 'products_export.xlsx')
        wb.save(file_path)

        self.stdout.write(
            self.style.SUCCESS(f'Successfully exported {products.count()} products to {file_path}')
        )
