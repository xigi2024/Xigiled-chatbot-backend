from django.core.management.base import BaseCommand
from Alexa.models import Product
from Alexa.views import INDOOR_SPECS, OUTDOOR_SPECS, RENTAL_SPECS
import openpyxl

class Command(BaseCommand):
    help = 'Export panel specs, guides, and specifications to Excel'

    def handle(self, *args, **options):
        wb = openpyxl.Workbook()

        # Indoor Specs
        ws_indoor = wb.active
        ws_indoor.title = 'Indoor Specs'
        headers = ['Model', 'Pixel Pitch', 'Module Resolution', 'LED Type', 'Brightness', 'Module Size', 'Scan Time', 'Driving Mode', 'IP Rating', 'Price per Sq.m', 'Price per Cabinet']
        ws_indoor.append(headers)
        for model, specs in INDOOR_SPECS.items():
            row = [
                model,
                specs.get('pixel_pitch', ''),
                ', '.join(specs.get('module_resolutions', [])),
                ', '.join(specs.get('led_types', [])),
                ', '.join(specs.get('brightness_options', [])),
                ', '.join(specs.get('module_sizes', [])),
                ', '.join(specs.get('scan_times', [])) if specs.get('scan_times') else '',
                ', '.join(specs.get('driving_modes', [])) if specs.get('driving_modes') else '',
                specs.get('ip_rating', ''),
                specs.get('price_per_sq_meter', ''),
                specs.get('price_per_cabinet', '')
            ]
            ws_indoor.append(row)

        # Outdoor Specs
        ws_outdoor = wb.create_sheet('Outdoor Specs')
        ws_outdoor.append(headers)
        for model, specs in OUTDOOR_SPECS.items():
            row = [
                model,
                specs.get('pixel_pitch', ''),
                ', '.join(specs.get('module_resolutions', [])),
                ', '.join(specs.get('led_types', [])),
                ', '.join(specs.get('brightness_options', [])),
                ', '.join(specs.get('module_sizes', [])),
                ', '.join(specs.get('scan_times', [])) if specs.get('scan_times') else '',
                ', '.join(specs.get('driving_modes', [])) if specs.get('driving_modes') else '',
                specs.get('ip_rating', ''),
                specs.get('price_per_sq_meter', ''),
                specs.get('price_per_cabinet', '')
            ]
            ws_outdoor.append(row)

        # Rental Specs
        ws_rental = wb.create_sheet('Rental Specs')
        rental_headers = ['Model', 'Pixel Pitch', 'Module Resolution', 'LED Type', 'Brightness', 'Module Size', 'Scan Time', 'IP Rating', 'Rental Price per Day', 'Rental Price per Week', 'Setup Fee', 'Durability', 'Availability']
        ws_rental.append(rental_headers)
        for model, specs in RENTAL_SPECS.items():
            row = [
                model,
                specs.get('pixel_pitch', ''),
                ', '.join(specs.get('module_resolutions', [])),
                ', '.join(specs.get('led_types', [])),
                ', '.join(specs.get('brightness_options', [])),
                ', '.join(specs.get('module_sizes', [])),
                ', '.join(specs.get('scan_times', [])) if specs.get('scan_times') else '',
                specs.get('ip_rating', ''),
                specs.get('rental_price_per_day', ''),
                specs.get('rental_price_per_week', ''),
                specs.get('setup_fee', ''),
                specs.get('durability', ''),
                specs.get('availability', '')
            ]
            ws_rental.append(row)

        # Guides and Specifications from Product model
        ws_guides = wb.create_sheet('Guides and Specifications')
        guide_headers = ['Name', 'Description', 'Price', 'Category', 'Guide Steps']
        ws_guides.append(guide_headers)
        products = Product.objects.all()
        for product in products:
            guide_steps = '\n'.join(product.guide_steps) if product.guide_steps else ''
            row = [
                product.name,
                product.description,
                product.price,
                product.category,
                guide_steps
            ]
            ws_guides.append(row)

        wb.save('panel_specs_export.xlsx')
        self.stdout.write(self.style.SUCCESS('Successfully exported panel specs, guides, and specifications to panel_specs_export.xlsx'))
