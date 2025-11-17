from django.urls import path
from .views import AlexaChatAPIView, AnalyticsAPIView, ChatDataAPIView, WelcomeAPIView, EnhancedWelcomeAPIView, CustomWelcomeAPIView
from django.http import HttpResponse
import os
import openpyxl

def export_products_view(request):
    # Set up Django environment
    import django
    from django.conf import settings
    if not settings.configured:
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'myassistant.settings')
        django.setup()

    from .models import Product

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = ['Name', 'Description', 'Price', 'Category', 'Guide Steps']
    ws.append(headers)

    # Fetch all products
    products = Product.objects.all()

    for product in products:
        # Combine guide_steps into a single string with line breaks
        guide_steps_str = '\n'.join(product.guide_steps) if product.guide_steps else ''

        # Append row data
        row = [
            product.name,
            product.description,
            product.price,
            product.category or '',
            guide_steps_str
        ]
        ws.append(row)

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)
    return response

def export_specs_view(request):
    # Hardcoded specs from views.py
    INDOOR_SPECS = {
        'P1.25mm': {
            'module_resolutions': ['256x128'],
            'led_types': ['SMD 3 in1'],
            'brightness_options': ['600-800'],
            'module_sizes': ['320x160'],
            'scan_times': ['1/64Scan'],
            'pixel_pitch': 'P1.25mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹250,000 – ₹320,000',
            'price_per_cabinet': '₹60,000 – ₹80,000'
        },
        'P2.5mm': {
            'module_resolutions': ['128x64'],
            'led_types': ['SMD 3 in1'],
            'brightness_options': ['800-1000'],
            'module_sizes': ['320x160'],
            'scan_times': ['1/32Scan'],
            'pixel_pitch': 'P2.5mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹95,000 – ₹125,000',
            'price_per_cabinet': '₹25,000 – ₹32,000'
        },
        'P3mm': {
            'module_resolutions': ['192x192'],
            'led_types': ['SMD 3 in1'],
            'brightness_options': ['900'],
            'module_sizes': ['320x160'],
            'scan_times': ['1/16Scan'],
            'pixel_pitch': 'P3mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹75,000 – ₹95,000',
            'price_per_cabinet': '₹20,000 – ₹25,000'
        },
        'P3.91mm': {
            'module_resolutions': ['64x64'],
            'led_types': ['SMD 3 in1'],
            'brightness_options': ['800-1000'],
            'module_sizes': ['250x250'],
            'scan_times': ['1/16Scan'],
            'pixel_pitch': 'P3.91mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹65,000 – ₹85,000',
            'price_per_cabinet': '₹18,000 – ₹22,000'
        },
        'P4.81mm': {
            'module_resolutions': ['52x52'],
            'led_types': ['SMD 3 in1'],
            'brightness_options': ['800-1000'],
            'module_sizes': ['250x250'],
            'scan_times': ['1/13Scan'],
            'pixel_pitch': 'P4.81mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹60,000 – ₹75,000',
            'price_per_cabinet': '₹16,000 – ₹20,000'
        },
        'P6mm': {
            'module_resolutions': ['32x16'],
            'led_types': ['White LED'],
            'brightness_options': ['800-1000'],
            'module_sizes': ['192x96'],
            'driving_modes': ['1/8Scan'],
            'pixel_pitch': 'P6mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹50,000 – ₹65,000',
            'price_per_cabinet': '₹14,000 – ₹17,000'
        },
        'P10mm': {
            'module_resolutions': ['32x16'],
            'led_types': ['White LED'],
            'brightness_options': ['800-1000'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/4Scan'],
            'pixel_pitch': 'P10mm',
            'ip_rating': 'IP30',
            'price_per_sq_meter': '₹40,000 – ₹55,000',
            'price_per_cabinet': '₹12,000 – ₹15,000'
        }
    }

    OUTDOOR_SPECS = {
        'P3.076mm': {
            'module_resolutions': ['104x52'],
            'led_types': ['SMD1415'],
            'brightness_options': ['>5000'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/13Scan'],
            'pixel_pitch': 'P3.076mm',
            'ip_rating': 'IP65'
        },
        'P4mm': {
            'module_resolutions': ['80x40'],
            'led_types': ['SMD1921'],
            'brightness_options': ['>5500'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/10Scan'],
            'pixel_pitch': 'P4mm',
            'ip_rating': 'IP67'
        },
        'P5mm': {
            'module_resolutions': ['64x32'],
            'led_types': ['SMD1921'],
            'brightness_options': ['>5200'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/8Scan'],
            'pixel_pitch': 'P5mm',
            'ip_rating': 'IP67',
            'price_per_sq_meter': '₹55,000 – ₹70,000',
            'price_per_cabinet': '₹15,000 – ₹18,000'
        },
        'P6.67mm': {
            'module_resolutions': ['48x24'],
            'led_types': ['SMD3535'],
            'brightness_options': ['>5500'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/6Scan'],
            'pixel_pitch': 'P6.67mm',
            'ip_rating': 'IP67',
            'price_per_sq_meter': '₹50,000 – ₹65,000',
            'price_per_cabinet': '₹14,000 – ₹17,000'
        },
        'P8mm': {
            'module_resolutions': ['40x20'],
            'led_types': ['SMD3535'],
            'brightness_options': ['>5800'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/5Scan'],
            'pixel_pitch': 'P8mm',
            'ip_rating': 'IP67'
        },
        'P10mm': {
            'module_resolutions': ['32x16'],
            'led_types': ['SMD3535'],
            'brightness_options': ['>6000'],
            'module_sizes': ['320x160'],
            'driving_modes': ['1/2Scan'],
            'pixel_pitch': 'P10mm',
            'ip_rating': 'IP67',
            'price_per_sq_meter': '₹40,000 – ₹55,000',
            'price_per_cabinet': '₹12,000 – ₹15,000'
        }
    }

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panel Specs"

    # Headers
    headers = [
        'Type', 'Model', 'Pixel Pitch', 'Module Resolution', 'LED Type',
        'Brightness', 'Module Size', 'Scan Time / Driving Mode', 'IP Rating',
        'Price per Sq.Meter', 'Price per Cabinet'
    ]
    ws.append(headers)

    # Export Indoor Specs
    for model, specs in INDOOR_SPECS.items():
        row = [
            'Indoor',
            model,
            specs.get('pixel_pitch', ''),
            ', '.join(specs.get('module_resolutions', [])),
            ', '.join(specs.get('led_types', [])),
            ', '.join(specs.get('brightness_options', [])),
            ', '.join(specs.get('module_sizes', [])),
            ', '.join(specs.get('scan_times', [])) or ', '.join(specs.get('driving_modes', [])),
            specs.get('ip_rating', ''),
            specs.get('price_per_sq_meter', ''),
            specs.get('price_per_cabinet', '')
        ]
        ws.append(row)

    # Export Outdoor Specs
    for model, specs in OUTDOOR_SPECS.items():
        row = [
            'Outdoor',
            model,
            specs.get('pixel_pitch', ''),
            ', '.join(specs.get('module_resolutions', [])),
            ', '.join(specs.get('led_types', [])),
            ', '.join(specs.get('brightness_options', [])),
            ', '.join(specs.get('module_sizes', [])),
            ', '.join(specs.get('scan_times', [])) or ', '.join(specs.get('driving_modes', [])),
            specs.get('ip_rating', ''),
            specs.get('price_per_sq_meter', ''),
            specs.get('price_per_cabinet', '')
        ]
        ws.append(row)

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=panel_specs.xlsx'
    wb.save(response)
    return response

def export_guides_view(request):
    # Set up Django environment
    import django
    from django.conf import settings
    if not settings.configured:
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'myassistant.settings')
        django.setup()

    from .models import Product

    # Hardcoded purpose recommendations from views.py
    PURPOSE_RECOMMENDATIONS = {
        'event hall': {
            'tips': [
                'Ensure high brightness for large viewing distances (typically 5000+ nits).',
                'Consider modular panels for flexible configuration in different venues.',
                'Recommend controllers with hot-swapping capability for quick repairs.',
                'Consider rental vs. purchase based on frequency of use.'
            ],
            'additional_accessories': ['Truss Systems', 'Rigging Hardware', 'Quick-Release Mounting Brackets'],
            'panel_recommendation': 'Outdoor panels (P3-P5mm) for high brightness; Indoor P2.5-P3.91mm for auditorium setups',
            'estimated_brightness': 'Minimum 5000 nits for distance viewing',
            'setup_steps': [
                '1. Assess the venue layout and determine optimal screen placement for maximum visibility from all seating areas.',
                '2. Ensure power supply and cabling can handle the display load; consider backup generators for critical events.',
                '3. Test brightness and color accuracy in the actual lighting conditions of the hall.',
                '4. Coordinate with event staff for rigging, safety checks, and emergency evacuation routes.',
                '5. Perform a full system test, including failover scenarios, before the event starts.',
                '6. Monitor temperature and ventilation to prevent overheating during long events.'
            ]
        },
        'studio': {
            'tips': [
                'Prioritize color accuracy and low input latency for live broadcasts.',
                'Choose smaller pixel pitch for detailed content visibility.',
                'Consider background lighting vs. main display based on studio size.',
                'Plan for thermal management in enclosed spaces.'
            ],
            'additional_accessories': ['Color Calibration Tools', 'Green Screen Backdrops', 'LED Color Management Software'],
            'panel_recommendation': 'Indoor P1.25-P2.5mm for high definition and color accuracy',
            'estimated_brightness': '600-800 nits for controlled indoor lighting',
            'setup_steps': [
                '1. Position panels to avoid reflections and ensure even lighting across the studio.',
                '2. Calibrate color temperature to match studio lighting and camera settings.',
                '3. Integrate with broadcast equipment for low-latency signal transmission.',
                '4. Install in a controlled environment to maintain consistent performance.',
                '5. Test with actual broadcast scenarios, including live feeds and recordings.',
                '6. Implement cooling systems if panels will be used for extended periods.'
            ]
        },
        'mall': {
            'tips': [
                'High visibility and durability essential for 24/7 operation.',
                'Integrate with existing mall signage systems and digital networks.',
                'Plan for remote content management across multiple displays.',
                'Consider energy consumption for continuous operation.'
            ],
            'additional_accessories': ['Digital Signage Software', 'Content Management Systems', 'Network Integration Kits'],
            'panel_recommendation': 'Indoor P2.5-P4mm or Outdoor P4-P5mm depending on location (indoor vs. outdoor signage)',
            'estimated_brightness': 'Indoor: 800-1000 nits; Outdoor: 5000+ nits',
            'setup_steps': [
                '1. Choose locations with high foot traffic and minimal obstructions for optimal visibility.',
                '2. Ensure panels are securely mounted to withstand public interaction.',
                "3. Integrate with mall's central content management system for unified control.",
                '4. Schedule maintenance windows during low-traffic hours.',
                '5. Test display performance under varying crowd sizes and lighting conditions.',
                '6. Implement energy-saving modes for off-peak hours.'
            ]
        },
        'outdoor stage': {
            'tips': [
                'Weather-resistant and high brightness essential (6000+ nits for daylight).',
                'Consider wind and rain protection; ensure proper drainage.',
                'Plan for thermal dissipation in hot climates.',
                'Implement redundant power supplies for critical events.'
            ],
            'additional_accessories': ['Weatherproof Enclosures', 'Ground Stakes', 'Lightning Protection Kits', 'Thermal Management Systems'],
            'panel_recommendation': 'Outdoor panels P4-P6.67mm with IP67 or higher rating',
            'estimated_brightness': 'Minimum 6000 nits for outdoor daytime visibility',
            'setup_steps': [
                '1. Select weatherproof panels and enclosures suitable for the local climate.',
                '2. Secure panels against wind and vibration using appropriate mounting hardware.',
                '3. Position screens to avoid direct sunlight glare and ensure visibility from all audience areas.',
                '4. Install lightning protection and grounding systems.',
                '5. Test under simulated weather conditions and perform acoustic checks if near speakers.',
                '6. Have backup power and quick-replacement panels on site for large events.'
            ]
        },
        'church': {
            'tips': [
                'Subtle, warm lighting creates ambiance without distraction.',
                'Consider acoustic considerations and vibration isolation.',
                'Plan for gradual brightness adjustment during services.',
                'Integrate with audio systems for synchronized content.'
            ],
            'additional_accessories': ['Dimming Controllers', 'Sound Integration Kits', 'Subtle Color Temperature Controls'],
            'panel_recommendation': 'Indoor P3-P4mm with warm color temperature support',
            'estimated_brightness': '400-600 nits for comfortable viewing in dim environments',
            'setup_steps': [
                '1. Position displays to complement the worship space without dominating the environment.',
                '2. Use dimming controls to adjust brightness for different parts of the service.',
                '3. Ensure panels are mounted securely to avoid vibrations from music or movement.',
                '4. Integrate with sound systems for synchronized multimedia presentations.',
                '5. Test visibility from all seating areas, considering ambient lighting.',
                '6. Schedule installations during off-service times to minimize disruption.'
            ]
        },
        'temple': {
            'tips': [
                'Respect cultural sensitivities with subtle, reverent lighting.',
                'Consider acoustic and vibrational impacts in sacred spaces.',
                'Plan for adjustable brightness to accommodate various ceremonies.',
                'Ensure durability for frequent use in communal settings.'
            ],
            'additional_accessories': ['Dimming Controllers', 'Vibration Isolation Mounts', 'Cultural Content Management Software'],
            'panel_recommendation': 'Indoor P3-P4mm with warm color temperature and low-noise operation',
            'estimated_brightness': '400-600 nits for serene viewing in traditional settings',
            'setup_steps': [
                '1. Consult with temple authorities to align with cultural and religious guidelines.',
                '2. Position displays discreetly to maintain the sanctity of the space.',
                '3. Use vibration-isolated mounts to prevent disturbances during rituals.',
                '4. Implement gradual dimming for transitions between ceremony phases.',
                '5. Test audio integration carefully to avoid interference with chants or music.',
                '6. Provide training for temple staff on content management and maintenance.'
            ]
        },
        'retail': {
            'tips': [
                'Eye-catching displays to attract customer attention.',
                'Frequent content updates and dynamic messaging.',
                'Consider compact installation spaces.',
                'Energy efficiency for extended operating hours.'
            ],
            'additional_accessories': ['Dynamic Content Software', 'Compact Mounting Solutions', 'Remote Content Management'],
            'panel_recommendation': 'Indoor P2.5-P4mm for retail storefronts',
            'estimated_brightness': '800-1200 nits for bright retail environments',
            'setup_steps': [
                '1. Install in high-visibility areas like windows or entryways.',
                '2. Ensure panels are eye-catching but not overwhelming in a shopping environment.',
                '3. Integrate with retail management systems for real-time content updates.',
                '4. Use energy-efficient modes during non-business hours.',
                '5. Test display performance with actual product imagery and promotions.',
                '6. Plan for easy access for maintenance without disrupting store operations.'
            ]
        },
        'manufacturing factory': {
            'tips': [
                'Durable panels resistant to dust, vibrations, and harsh environments.',
                'High brightness for visibility in well-lit industrial spaces.',
                'Consider safety signage and process monitoring displays.',
                'Ensure compliance with industrial safety standards.'
            ],
            'additional_accessories': ['Rugged Enclosures', 'Vibration Dampeners', 'Industrial Power Supplies', 'Safety Integration Kits'],
            'panel_recommendation': 'Indoor P3-P5mm with high IP rating for dust and moisture resistance',
            'estimated_brightness': '1000-1500 nits for industrial lighting conditions',
            'setup_steps': [
                '1. Assess factory layout for optimal display placement without obstructing workflows.',
                '2. Choose panels with high IP ratings to withstand dust and occasional moisture.',
                '3. Secure mounts to handle vibrations from machinery.',
                '4. Integrate with factory control systems for real-time data display.',
                '5. Ensure displays comply with workplace safety regulations.',
                '6. Schedule maintenance during off-hours to minimize production downtime.'
            ]
        },
        'manufacturer': {
            'tips': [
                'Durable panels resistant to dust, vibrations, and harsh environments.',
                'High brightness for visibility in well-lit industrial spaces.',
                'Consider safety signage and process monitoring displays.',
                'Ensure compliance with industrial safety standards.'
            ],
            'additional_accessories': ['Rugged Enclosures', 'Vibration Dampeners', 'Industrial Power Supplies', 'Safety Integration Kits'],
            'panel_recommendation': 'Indoor P3-P5mm with high IP rating for dust and moisture resistance',
            'estimated_brightness': '1000-1500 nits for industrial lighting conditions',
            'setup_steps': [
                '1. Assess factory layout for optimal display placement without obstructing workflows.',
                '2. Choose panels with high IP ratings to withstand dust and occasional moisture.',
                '3. Secure mounts to handle vibrations from machinery.',
                '4. Integrate with factory control systems for real-time data display.',
                '5. Ensure displays comply with workplace safety regulations.',
                '6. Schedule maintenance during off-hours to minimize production downtime.'
            ]
        },
        'default': {
            'tips': ['Choose based on viewing distance, brightness requirements, and environment.'],
            'additional_accessories': [],
            'panel_recommendation': 'Consult with our team for personalized recommendations',
            'estimated_brightness': 'Depends on installation location',
            'setup_steps': [
                '1. Assess the specific requirements of your location and intended use.',
                '2. Consult with our technical team for tailored recommendations.',
                '3. Plan the installation with safety and accessibility in mind.',
                '4. Test all components in the actual environment before final setup.',
                '5. Schedule regular maintenance to ensure long-term performance.'
            ]
        }
    }

    # Create Excel file for guides
    wb = openpyxl.Workbook()
    
    # Purpose Guides Sheet
    ws_purpose = wb.active
    ws_purpose.title = "Purpose Guides"
    
    # Headers for Purpose Guides
    purpose_headers = ['Purpose', 'Panel Recommendation', 'Estimated Brightness', 'Key Considerations', 'Additional Accessories', 'Setup Steps']
    ws_purpose.append(purpose_headers)
    
    # Export Purpose Recommendations
    for purpose, recs in PURPOSE_RECOMMENDATIONS.items():
        if purpose != 'default':
            tips_str = '\n'.join(recs.get('tips', []))
            accessories_str = '\n'.join(recs.get('additional_accessories', []))
            setup_str = '\n'.join(recs.get('setup_steps', []))
            
            row = [
                purpose.title(),
                recs.get('panel_recommendation', ''),
                recs.get('estimated_brightness', ''),
                tips_str,
                accessories_str,
                setup_str
            ]
            ws_purpose.append(row)
    
    # Panel Guides Sheet (from Product model guide_steps)
    ws_panel = wb.create_sheet(title="Panel Guides")
    
    # Headers for Panel Guides
    panel_headers = ['Panel Name', 'Guide Steps']
    ws_panel.append(panel_headers)
    
    # Fetch all products and their guide_steps
    products = Product.objects.all()
    for product in products:
        guide_steps_str = '\n'.join(product.guide_steps) if product.guide_steps else ''
        row = [
            product.name,
            guide_steps_str
        ]
        ws_panel.append(row)

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=guides.xlsx'
    wb.save(response)
    return response

urlpatterns = [
    path('', AlexaChatAPIView.as_view(), name='alexa_chat_api'),
    path('analytics/', AnalyticsAPIView.as_view(), name='analytics_api'),
    path('chat-data/', ChatDataAPIView.as_view(), name='chat_data_api'),
    path('welcome/', WelcomeAPIView.as_view(), name='welcome_api'),
    path('enhanced-welcome/', EnhancedWelcomeAPIView.as_view(), name='enhanced_welcome_api'),
    path('custom-welcome/', CustomWelcomeAPIView.as_view(), name='custom_welcome_api'),
    path('export-products/', export_products_view, name='export_products'),
    path('export-specs/', export_specs_view, name='export_specs'),
    path('export-guides/', export_guides_view, name='export_guides'),
]
